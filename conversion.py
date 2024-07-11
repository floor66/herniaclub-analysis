from openpyxl import load_workbook

workbook = load_workbook("ONLY HEADERS.xlsx")

conversion_sheet = workbook["Conversion"]
cs = conversion_sheet

data_sheet = workbook["Raw"]
ds = data_sheet

# Constants
COL_HC_VARIABLE = 0
COL_STATS_VARIABLE = 1
COL_VARIABLE_DEF = 2
COL_VARIABLE_COD_FR = 3
COL_ORIG = 4
COL_CONV = 5
COL_LABEL = 6

NUM_COLS_DATA = 191
NUM_COLS_CONV = 7

CONV_MAX_ROW = 456

V_TYPE_CAT = "V_TYPE_CAT"
V_TYPE_COMPOUND = "V_TYPE_COMPOUND"
V_TYPE_CONT = "V_TYPE_CONT"
V_TYPE_CONT_SPEC = "V_TYPE_CONT_SPEC"
V_TYPE_MISSING = "V_TYPE_MISSING"
V_TYPE_OTHER = "V_TYPE_OTHER"

MISSING = 999

def determine_variable_type(row):
    if row[COL_STATS_VARIABLE].value is not None:
        if row[COL_STATS_VARIABLE].value[-1] == "_":
            return V_TYPE_COMPOUND

    if row[COL_VARIABLE_COD_FR].value == "Missing":
        return V_TYPE_MISSING

    if row[COL_VARIABLE_COD_FR].value == "Continuous":
        return V_TYPE_CONT

    if row[COL_VARIABLE_COD_FR].value == "Continuous special":
        return V_TYPE_CONT_SPEC

    if row[COL_VARIABLE_COD_FR].value is None or row[COL_VARIABLE_COD_FR].value == "Date":
        return V_TYPE_OTHER

    return V_TYPE_CAT

def gather_variables():
    variables = []
    tmp_variable = None

    for row in cs.iter_rows(min_row=2, max_row=CONV_MAX_ROW, max_col=NUM_COLS_CONV):
        rownr = row[0].row

        if row[COL_HC_VARIABLE].value is not None:
            if tmp_variable is not None:
                # print("Finished variable: %s" % tmp_variable["hc_name"])
                tmp_variable["end_row"] = rownr - 1
                variables.append(tmp_variable)

            # print("New variable: %s" % row[COL_HC_VARIABLE].value)
            tmp_variable = {
                "hc_name": row[COL_HC_VARIABLE].value,
                "type": determine_variable_type(row),
                "start_row": rownr
            }

    # print("Finished variable: %s" % tmp_variable["hc_name"])
    tmp_variable["end_row"] = rownr - 1
    variables.append(tmp_variable)

    return variables

def generate_spss_variable_syntax(var):
    syntax = "* Start generated SPSS syntax for Hernia Club variable '%s'.\n" % var["hc_name"]

    if var["type"] == V_TYPE_CAT:
        for row in cs.iter_rows(min_row=var["start_row"], max_row=var["end_row"], max_col=NUM_COLS_CONV):
            rownr = row[0].row

            if rownr == var["start_row"]:
                syntax += "COMPUTE %s = 999.\n" % row[COL_STATS_VARIABLE].value
                syntax += "MISSING VALUES %s(999).\n" % row[COL_STATS_VARIABLE].value
                syntax += "VARIABLE LABELS\n"
                syntax += "    %s\n" % row[COL_STATS_VARIABLE].value
                syntax += "    '%s'.\n" % row[COL_VARIABLE_DEF].value
                syntax += "VALUE LABELS\n"
                syntax += "    %s\n" % row[COL_STATS_VARIABLE].value

            syntax += "    %i '%s'" % (row[COL_CONV].value, row[COL_LABEL].value)

            if rownr == var["end_row"]:
                syntax += "."

            syntax += "\n"
    elif var["type"] == V_TYPE_COMPOUND:
        # (i) In some cases, there is a specific answer for the value '0' (e.g. 0 = Non)
        # (ii) In most cases, 0 will default to 'sans objet' which means not applicable
        # If we dont define an exception like in (i), add a _NOT_APPLICABLE parameter automatically
        zero_is_defined = False

        for row in cs.iter_rows(min_row=var["start_row"], max_row=var["end_row"], max_col=NUM_COLS_CONV):
            rownr = row[0].row

            if rownr == var["start_row"]:
                variable_name_prefix = row[COL_STATS_VARIABLE].value
                variable_label_prefix = row[COL_VARIABLE_DEF].value +": "
            else:
                # All following rows will describe a yes/no variable
                if row[COL_ORIG].value == 0:
                    zero_is_defined = True

                variable_name = variable_name_prefix + row[COL_STATS_VARIABLE].value
                variable_label = variable_label_prefix + row[COL_LABEL].value
                
                syntax += "COMPUTE %s = 999.\n" % variable_name
                syntax += "MISSING VALUES %s(999).\n" % variable_name
                syntax += "VARIABLE LABELS\n"
                syntax += "    %s\n" % variable_name
                syntax += "    '%s'.\n" % variable_label
                syntax += "VALUE LABELS\n"
                syntax += "    1 'Yes'\n"
                syntax += "    0 'No'.\n"
            
        if not zero_is_defined:
            variable_name = variable_name_prefix + "NOT_APPLICABLE"
            variable_label = variable_label_prefix + "Not applicable"

            syntax += "COMPUTE %s = 999.\n" % variable_name
            syntax += "MISSING VALUES %s(999).\n" % variable_name
            syntax += "VARIABLE LABELS\n"
            syntax += "    %s\n" % variable_name
            syntax += "    '%s'.\n" % variable_label
            syntax += "VALUE LABELS\n"
            syntax += "    1 'Yes'\n"
            syntax += "    0 'No'.\n"
    elif var["type"] == V_TYPE_CONT or var["type"] == V_TYPE_CONT_SPEC:
        # Continunous variables are a single row with a number variable
        row = cs[var["start_row"]]

        syntax += "COMPUTE %s = 999.\n" % row[COL_STATS_VARIABLE].value
        syntax += "VARIABLE LABELS\n"
        syntax += "    %s\n" % row[COL_STATS_VARIABLE].value
        syntax += "    '%s'.\n" % row[COL_VARIABLE_DEF].value
    elif var["type"] == V_TYPE_OTHER:
        # 'Other' variables are a single row with a string variable
        row = cs[var["start_row"]]

        syntax += "STRING %s(a128).\n" % row[COL_STATS_VARIABLE].value
        syntax += "VARIABLE LABELS\n"
        syntax += "    %s\n" % row[COL_STATS_VARIABLE].value
        syntax += "    '%s'.\n" % row[COL_VARIABLE_DEF].value
    elif var["type"] == V_TYPE_MISSING:
        return None

    syntax += "EXECUTE.\n"
    syntax += "* End syntax for '%s'.\n" % var["hc_name"]
    return syntax

def convert(var, val):
    if var["type"] != V_TYPE_COMPOUND and val is None:
        row = cs[var["start_row"]]

        return {
            row[COL_STATS_VARIABLE].value: MISSING
        }

    if var["type"] == V_TYPE_CONT:
        row = cs[var["start_row"]]

        return {
            row[COL_STATS_VARIABLE].value: val
        }
    elif var["type"] == V_TYPE_MISSING:
        row = cs[var["start_row"]]

        return {
            row[COL_STATS_VARIABLE].value: MISSING
        }
    elif var["type"] == V_TYPE_OTHER:
        row = cs[var["start_row"]]

        return {
            row[COL_STATS_VARIABLE].value: val
        }
    elif var["type"] == V_TYPE_CONT_SPEC:
        abc = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m",
               "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z",
               "aa", "ab", "ac", "ad", "ae", "af", "ag", "ah", "ai", "aj", "ak", "al", "am",
               "an", "ao", "ap", "aq", "ar", "as", "at", "au", "av", "aw", "ax", "ay", "az"]
        
        row = cs[var["start_row"]]

        return {
            row[COL_STATS_VARIABLE].value: abc.index(val)
        }
    elif var["type"] == V_TYPE_CAT:
        conversion = {}

        for row in cs.iter_rows(min_row=var["start_row"], max_row=var["end_row"], max_col=NUM_COLS_CONV):
            if row[0].row == var["start_row"]:
                variable_name = row[COL_STATS_VARIABLE].value
                
            conversion[row[COL_ORIG].value] = row[COL_CONV].value
        
        return {
            variable_name: conversion[val]
        }
    elif var["type"] == V_TYPE_COMPOUND:
        zero_is_defined = False
        conversion = {}
        result = {}

        for row in cs.iter_rows(min_row=var["start_row"], max_row=var["end_row"], max_col=NUM_COLS_CONV):
            if row[0].row == var["start_row"]:
                variable_name_prefix = row[COL_STATS_VARIABLE].value
            else:
                if row[COL_ORIG].value == 0:
                    zero_is_defined = True
                
                conversion[row[COL_ORIG].value] = variable_name_prefix + row[COL_STATS_VARIABLE].value
                result[variable_name_prefix + row[COL_STATS_VARIABLE].value] = 0
        
        if not zero_is_defined:
            conversion[0] = variable_name_prefix + "NOT_APPLICABLE"

        if val is None:
            for key in result.keys():
                result[key] = MISSING
        elif val == 0:
            result[conversion[0]] = 1
        else:
            # Do our sum deconstruction thing
            powers_of_two = list(conversion.keys())
            powers_of_two.sort()
            powers_of_two.reverse()

            for number in powers_of_two:
                if number == 0:
                    continue

                if val >= number:
                    result[conversion[number]] = 1
                    val -= number

        return result

# Skip headers that are empty
# Variables are enumerated 1..n WITHOUT empty headers
skip_columns = []
for cell in ds[1]:
    if cell.value is None:
        skip_columns.append(cell.column)

# Patient data conversion
def process_patients(row_start, row_end):
    patient = {}

    for row in ds.iter_rows(min_row=row_start, max_row=row_end, max_col=NUM_COLS_DATA):
        print([cell.value for cell in row])
        print("")

        i = 0
        for cell in row:
            if cell.column in skip_columns:
                continue

            # We havent converted all variables yet
            try:
                var = variables[i]
            except IndexError:
                break

            converted = convert(var, cell.value)

            for key in converted.keys():
                patient[key] = converted[key]

            i += 1

    return patient

# SPSS syntax for variables
# for i, var in enumerate(variables):
#     print(i, var)
#     print(generate_spss_variable_syntax(var))

variables = gather_variables() or {}
print(process_patients(3, 3))
